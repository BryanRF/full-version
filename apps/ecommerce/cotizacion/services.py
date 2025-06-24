# apps/ecommerce/cotizacion/services.py
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import date, timedelta
import logging
from apps.ecommerce.requirements.services import RequirementExcelGenerator
import pandas as pd
from decimal import Decimal, InvalidOperation
import logging

logger = logging.getLogger(__name__)

class CotizacionResponseProcessor:
    """Procesador automático de respuestas de cotización desde Excel"""
    
    def __init__(self, envio_cotizacion):
        self.envio = envio_cotizacion
        self.errors = []
        self.warnings = []
        self.processed_items = []
        
        # Mapeo oficial de columnas según tu template
        self.OFFICIAL_COLUMNS = {
            'CÓDIGO': 0,           # Columna A
            'PRODUCTO': 1,         # Columna B  
            'CATEGORÍA': 2,        # Columna C
            'CANT. SOLICITADA': 3, # Columna D
            'UNIDAD': 4,           # Columna E
            'CANT. DISPONIBLE': 5, # Columna F
            'PRECIO U.': 6,        # Columna G - CRÍTICO
            'PRECIO TOTAL': 7,     # Columna H
            'OBSERVACIONES': 8     # Columna I
        }
        
    def _validate_excel_from_file(self, file_obj):
        """Validar archivo Excel directamente desde el objeto file"""
        print(f"\n=== INICIANDO VALIDACIÓN DE ARCHIVO EN MEMORIA ===")
        
        try:
            # Resetear posición del archivo
            file_obj.seek(0)
            
            print("\n[1/3] Leyendo archivo Excel desde memoria...")
            df = pd.read_excel(file_obj, engine='openpyxl')
            print(f"✔ Archivo leído correctamente. Filas: {len(df)}, Columnas: {len(df.columns)}")
            
            # NO resetear aquí - mantener posición para que Django pueda leer después
            
            errors = []
            warnings = []
            
            print(f"\n[2/3] Validando estructura...")
            if len(df.columns) < len(self.OFFICIAL_COLUMNS):
                error_msg = f"El archivo debe tener al menos {len(self.OFFICIAL_COLUMNS)} columnas"
                errors.append(error_msg)
                return {
                    'valid': False,
                    'errors': errors,
                    'warnings': warnings,
                    'dataframe': None
                }
            
            if len(df) == 0:
                errors.append("El archivo Excel está vacío")
                return {
                    'valid': False,
                    'errors': errors,
                    'warnings': warnings,
                    'dataframe': None
                }
            
            print(f"\n[3/3] Validando contenido...")
            # Validar códigos y precios
            codigo_col_pos = self.OFFICIAL_COLUMNS['CÓDIGO']
            precio_col_pos = self.OFFICIAL_COLUMNS['PRECIO U.']
            
            codigo_vals = df.iloc[:, codigo_col_pos].dropna()
            precio_vals = df.iloc[:, precio_col_pos].dropna()
            
            if codigo_vals.empty:
                warnings.append("La columna CÓDIGO está completamente vacía")
            
            if precio_vals.empty:
                warnings.append("La columna PRECIO U. está completamente vacía")
            
            # Validar códigos de productos
            if not codigo_vals.empty:
                from apps.ecommerce.products.models import Product
                valid_codes = set(Product.objects.values_list('code', flat=True))
                codes_in_file = set(str(code).strip().upper() for code in codigo_vals if str(code).strip().upper() != 'NAN')
                matching_codes = codes_in_file & valid_codes
                
                if not matching_codes:
                    errors.append("No se encontraron códigos de productos válidos en el archivo")
                elif len(matching_codes) < len(codes_in_file) / 2:
                    warnings.append(f"Solo {len(matching_codes)} de {len(codes_in_file)} códigos son válidos")
            
            # Validar precios
            if not precio_vals.empty:
                precios_validos = sum(1 for precio in precio_vals if self._parse_price_safe(precio) is not None)
                if precios_validos == 0:
                    errors.append("No se encontraron precios válidos en el archivo")
                elif precios_validos < len(precio_vals) / 2:
                    warnings.append(f"Solo {precios_validos} de {len(precio_vals)} precios son válidos")
            
            is_valid = len(errors) == 0
            print(f"\n=== RESULTADO VALIDACIÓN: {'VÁLIDO' if is_valid else 'INVÁLIDO'} ===")
            
            return {
                'valid': is_valid,
                'errors': errors,
                'warnings': warnings,
                'dataframe': df if is_valid else None
            }
            
        except Exception as e:
            print(f"\n!!! EXCEPCIÓN DURANTE VALIDACIÓN: {str(e)}")
            return {
                'valid': False,
                'errors': [f"Error leyendo archivo: {str(e)}"],
                'warnings': [],
                'dataframe': None
            }
    def process_excel_from_memory(self, file_obj):
        """Procesar archivo Excel directamente desde memoria"""
        from datetime import datetime
        
        print("\n" + "="*50)
        print("PROCESAMIENTO DE ARCHIVO EXCEL EN MEMORIA")
        print("="*50)
        
        start_time = datetime.now()
        
        try:
            # Validar archivo
            validation_result = self._validate_excel_from_file(file_obj)
            
            if not validation_result['valid']:
                print("\n✖ Validación fallida:")
                for error in validation_result['errors']:
                    print(f"  - {error}")
                return {
                    'success': False,
                    'error': 'Archivo no válido',
                    'validation_errors': validation_result['errors']
                }
            
            df = validation_result['dataframe']
            
            # Crear respuesta de cotización
            print("\nCreando respuesta de cotización...")
            respuesta = self._create_cotizacion_response()
            print(f"✔ Respuesta creada con ID: {respuesta.id}")
            
            # Procesar productos
            print("\nProcesando productos...")
            success_count = self._process_products_fixed(df, respuesta)
            
            end_time = datetime.now()
            processing_time = end_time - start_time
            
            print(f"\nTiempo total: {processing_time}")
            
            if success_count > 0:
                print("\n✔ Procesamiento exitoso")
                return {
                    'success': True,
                    'respuesta_id': respuesta.id,
                    'processed_items': success_count,
                    'total_items': len(df),
                    'errors': self.errors,
                    'warnings': self.warnings
                }
            else:
                respuesta.delete()
                return {
                    'success': False,
                    'error': 'No se pudo procesar ningún producto',
                    'details': self.errors
                }
                
        except Exception as ex:
            print(f"\n!!! ERROR: {str(ex)}")
            return {
                'success': False,
                'error': f'Error inesperado: {str(ex)}'
            }
    def process_excel_file_with_logging_direct(self, file_path):
        """Procesar archivo Excel con ruta directa"""
        from datetime import datetime
        import os
        
        print("\n" + "="*50)
        print("INICIANDO PROCESAMIENTO DIRECTO DE ARCHIVO EXCEL")
        print("="*50)
        
        start_time = datetime.now()
        
        try:
            # Verificar que el archivo existe
            if not os.path.exists(file_path):
                error_msg = f'Archivo no existe en la ruta: {file_path}'
                print(f"✖ {error_msg}")
                return {'success': False, 'error': error_msg}
            
            print(f"\nArchivo a procesar: {file_path}")
            print(f"Archivo existe: {os.path.exists(file_path)}")
            print(f"Tamaño del archivo: {os.path.getsize(file_path)} bytes")
            
            validation_result = self._validate_excel_file(file_path)
            
            if not validation_result['valid']:
                print("\n✖ Validación fallida:")
                for error in validation_result['errors']:
                    print(f"  - {error}")
                return {
                    'success': False, 
                    'error': 'Archivo no válido',
                    'details': validation_result['errors'],
                    'validation_errors': validation_result['errors']
                }
            
            # Si la validación es exitosa, procesar
            df = validation_result['dataframe']
            
            # Crear respuesta de cotización
            print("\nCreando respuesta de cotización...")
            respuesta = self._create_cotizacion_response()
            print(f"✔ Respuesta creada con ID: {respuesta.id}")
            
            # Procesar productos
            print("\nProcesando productos...")
            success_count = self._process_products_fixed(df, respuesta)
            
            end_time = datetime.now()
            processing_time = end_time - start_time
            
            print(f"\nTiempo total de procesamiento: {processing_time}")
            
            if success_count > 0:
                print("\n✔ Procesamiento exitoso")
                return {
                    'success': True,
                    'respuesta_id': respuesta.id,
                    'processed_items': success_count,
                    'total_items': len(df),
                    'errors': self.errors,
                    'warnings': self.warnings,
                    'processing_time': str(processing_time)
                }
            else:
                # Si no se procesó ningún producto, eliminar la respuesta creada
                print("\n✖ No se procesaron productos")
                respuesta.delete()
                return {
                    'success': False,
                    'error': 'No se pudo procesar ningún producto',
                    'details': self.errors
                }
                
        except Exception as ex:
            end_time = datetime.now()
            print(f"\n!!! ERROR INESPERADO: {str(ex)}")
            logger.error(f"Error procesando archivo de cotización: {ex}")
            return {
                'success': False,
                'error': f'Error inesperado: {str(ex)}'
            }

    def _create_cotizacion_response(self):
        """Crear respuesta de cotización automática"""
        from .models import RespuestaCotizacion
        
        # Verificar si ya existe una respuesta
        try:
            existing_response = RespuestaCotizacion.objects.filter(envio=self.envio).first()
            if existing_response:
                print(f"✖ Ya existe respuesta con ID: {existing_response.id}")
                existing_response.delete()
                print("✔ Respuesta anterior eliminada")
        except Exception as e:
            print(f"⚠ Error verificando respuesta anterior: {e}")
        
        return RespuestaCotizacion.objects.create(
            envio=self.envio,
            terminos_pago="Por definir",
            tiempo_entrega="Según especificaciones",
            observaciones="Respuesta procesada automáticamente desde Excel",
            procesado_automaticamente=True
        )

        def _validate_excel_from_file(self, file_obj):
            """Validar archivo Excel directamente desde el objeto file"""
            print(f"\n=== INICIANDO VALIDACIÓN DE ARCHIVO EN MEMORIA ===")
            
            try:
                # Resetear posición del archivo
                file_obj.seek(0)
                
                print("\n[1/3] Leyendo archivo Excel desde memoria...")
                df = pd.read_excel(file_obj, engine='openpyxl')
                print(f"✔ Archivo leído correctamente. Filas: {len(df)}, Columnas: {len(df.columns)}")
                
                # NO resetear aquí - mantener posición para que Django pueda leer después
                
                errors = []
                warnings = []
                
                print(f"\n[2/3] Validando estructura...")
                if len(df.columns) < len(self.OFFICIAL_COLUMNS):
                    error_msg = f"El archivo debe tener al menos {len(self.OFFICIAL_COLUMNS)} columnas"
                    errors.append(error_msg)
                    return {
                        'valid': False,
                        'errors': errors,
                        'warnings': warnings,
                        'dataframe': None
                    }
                
                if len(df) == 0:
                    errors.append("El archivo Excel está vacío")
                    return {
                        'valid': False,
                        'errors': errors,
                        'warnings': warnings,
                        'dataframe': None
                    }
                
                print(f"\n[3/3] Validando contenido...")
                # Validar códigos y precios
                codigo_col_pos = self.OFFICIAL_COLUMNS['CÓDIGO']
                precio_col_pos = self.OFFICIAL_COLUMNS['PRECIO U.']
                
                codigo_vals = df.iloc[:, codigo_col_pos].dropna()
                precio_vals = df.iloc[:, precio_col_pos].dropna()
                
                if codigo_vals.empty:
                    warnings.append("La columna CÓDIGO está completamente vacía")
                
                if precio_vals.empty:
                    warnings.append("La columna PRECIO U. está completamente vacía")
                
                # Validar códigos de productos
                if not codigo_vals.empty:
                    from apps.ecommerce.products.models import Product
                    valid_codes = set(Product.objects.values_list('code', flat=True))
                    codes_in_file = set(str(code).strip().upper() for code in codigo_vals if str(code).strip().upper() != 'NAN')
                    matching_codes = codes_in_file & valid_codes
                    
                    if not matching_codes:
                        errors.append("No se encontraron códigos de productos válidos en el archivo")
                    elif len(matching_codes) < len(codes_in_file) / 2:
                        warnings.append(f"Solo {len(matching_codes)} de {len(codes_in_file)} códigos son válidos")
                
                # Validar precios
                if not precio_vals.empty:
                    precios_validos = sum(1 for precio in precio_vals if self._parse_price_safe(precio) is not None)
                    if precios_validos == 0:
                        errors.append("No se encontraron precios válidos en el archivo")
                    elif precios_validos < len(precio_vals) / 2:
                        warnings.append(f"Solo {precios_validos} de {len(precio_vals)} precios son válidos")
                
                is_valid = len(errors) == 0
                print(f"\n=== RESULTADO VALIDACIÓN: {'VÁLIDO' if is_valid else 'INVÁLIDO'} ===")
                
                return {
                    'valid': is_valid,
                    'errors': errors,
                    'warnings': warnings,
                    'dataframe': df if is_valid else None
                }
                
            except Exception as e:
                print(f"\n!!! EXCEPCIÓN DURANTE VALIDACIÓN: {str(e)}")
                return {
                    'valid': False,
                    'errors': [f"Error leyendo archivo: {str(e)}"],
                    'warnings': [],
                    'dataframe': None
                }
            
    def _validate_excel_file(self, file_path):
        """Validar archivo Excel con estructura oficial"""
        print(f"\n=== INICIANDO VALIDACIÓN DE ARCHIVO ===")
        print(f"Ruta del archivo: {file_path}")
        
        try:
            # Leer archivo Excel
            print("\n[1/5] Intentando leer archivo Excel...")
            df = self._read_excel_file(file_path)
            
            if df is None:
                print("!!! ERROR: No se pudo leer el archivo Excel")
                return {
                    'valid': False,
                    'errors': ['No se pudo leer el archivo Excel'],
                    'warnings': [],
                    'suggestions': [],
                    'dataframe': None
                }
            
            print(f"✔ Archivo leído correctamente. Filas: {len(df)}, Columnas: {len(df.columns)}")
            print(f"Columnas detectadas: {list(df.columns)}")
            
            errors = []
            warnings = []
            suggestions = []
            
            # Validar que tenga suficientes columnas
            print(f"\n[2/5] Validando número de columnas...")
            print(f"Columnas esperadas: {len(self.OFFICIAL_COLUMNS)}, Columnas encontradas: {len(df.columns)}")
            
            if len(df.columns) < len(self.OFFICIAL_COLUMNS):
                error_msg = f"El archivo debe tener al menos {len(self.OFFICIAL_COLUMNS)} columnas"
                print(f"!!! ERROR: {error_msg}")
                errors.append(error_msg)
                return {
                    'valid': False,
                    'errors': errors,
                    'warnings': warnings,
                    'suggestions': ['Asegúrese de usar el template oficial'],
                    'dataframe': None
                }
            
            # Validar estructura de encabezados (primera fila)
            print(f"\n[3/5] Validando contenido del archivo...")
            print(f"Filas totales: {len(df)}")
            
            if len(df) == 0:
                error_msg = "El archivo Excel está vacío"
                print(f"!!! ERROR: {error_msg}")
                errors.append(error_msg)
                return {
                    'valid': False,
                    'errors': errors,
                    'warnings': warnings,
                    'suggestions': [],
                    'dataframe': None
                }
            
            # Verificar que hay datos más allá de los encabezados
            if len(df) < 2:
                warning_msg = "El archivo solo contiene encabezados, no hay productos para procesar"
                print(f"⚠ ADVERTENCIA: {warning_msg}")
                warnings.append(warning_msg)
            
            # Validar columnas críticas
            codigo_col_pos = self.OFFICIAL_COLUMNS['CÓDIGO']
            precio_col_pos = self.OFFICIAL_COLUMNS['PRECIO U.']
            
            print(f"\n[4/5] Validando columnas críticas...")
            print(f"Posición columna CÓDIGO: {codigo_col_pos}")
            print(f"Posición columna PRECIO U.: {precio_col_pos}")
            
            # Verificar que las columnas críticas no estén completamente vacías
            codigo_vals = df.iloc[:, codigo_col_pos].dropna()
            precio_vals = df.iloc[:, precio_col_pos].dropna()
            
            print(f"Valores en columna CÓDIGO: {len(codigo_vals)}/{len(df)}")
            print(f"Valores en columna PRECIO U.: {len(precio_vals)}/{len(df)}")
            
            if codigo_vals.empty:
                warning_msg = "La columna CÓDIGO está completamente vacía"
                print(f"⚠ ADVERTENCIA: {warning_msg}")
                warnings.append(warning_msg)
            
            if precio_vals.empty:
                warning_msg = "La columna PRECIO U. está completamente vacía"
                print(f"⚠ ADVERTENCIA: {warning_msg}")
                warnings.append(warning_msg)
            
            # Validar al menos algunos códigos de productos válidos
            if not codigo_vals.empty:
                print(f"\n[5/5] Validando códigos de productos...")
                from apps.ecommerce.products.models import Product
                valid_codes = set(Product.objects.values_list('code', flat=True))
                print(f"Códigos válidos en sistema: {len(valid_codes)}")
                
                codes_in_file = set(str(code).strip().upper() for code in codigo_vals if str(code).strip().upper() != 'NAN')
                print(f"Códigos únicos en archivo: {len(codes_in_file)}")
                
                matching_codes = codes_in_file & valid_codes
                print(f"Códigos coincidentes: {len(matching_codes)}")
                
                if not matching_codes:
                    error_msg = "No se encontraron códigos de productos válidos en el archivo"
                    print(f"!!! ERROR: {error_msg}")
                    errors.append(error_msg)
                    suggestions.append("Verifique que los códigos coincidan exactamente con los del sistema")
                elif len(matching_codes) < len(codes_in_file) / 2:
                    warning_msg = f"Solo {len(matching_codes)} de {len(codes_in_file)} códigos son válidos"
                    print(f"⚠ ADVERTENCIA: {warning_msg}")
                    warnings.append(warning_msg)
            
            # Validar formato de precios
            if not precio_vals.empty:
                print("\nValidando formatos de precios...")
                precios_validos = 0
                for i, precio in enumerate(precio_vals):
                    parsed = self._parse_price_safe(precio)
                    if parsed is not None:
                        precios_validos += 1
                    else:
                        print(f"  Precio inválido en fila {i+2}: {precio}")
                
                print(f"Precios válidos: {precios_validos}/{len(precio_vals)}")
                
                if precios_validos == 0:
                    error_msg = "No se encontraron precios válidos en el archivo"
                    print(f"!!! ERROR: {error_msg}")
                    errors.append(error_msg)
                    suggestions.append("Los precios deben ser números positivos (ej: 15.50)")
                elif precios_validos < len(precio_vals) / 2:
                    warning_msg = f"Solo {precios_validos} de {len(precio_vals)} precios son válidos"
                    print(f"⚠ ADVERTENCIA: {warning_msg}")
                    warnings.append(warning_msg)
            
            # Determinar si es válido
            is_valid = len(errors) == 0
            print(f"\n=== RESULTADO VALIDACIÓN: {'VÁLIDO' if is_valid else 'INVÁLIDO'} ===")
            print(f"Errores: {errors}")
            print(f"Advertencias: {warnings}")
            
            return {
                'valid': is_valid,
                'errors': errors,
                'warnings': warnings,
                'suggestions': suggestions,
                'dataframe': df if is_valid else None
            }
            
        except Exception as e:
            print(f"\n!!! EXCEPCIÓN DURANTE VALIDACIÓN: {str(e)}")
            return {
                'valid': False,
                'errors': [f"Error leyendo archivo: {str(e)}"],
                'warnings': [],
                'suggestions': ['Verifique que el archivo no esté corrupto y sea un Excel válido'],
                'dataframe': None
            }

    def _read_excel_file(self, file_path):
        """Leer archivo Excel usando pandas"""
        print(f"\nIntentando leer archivo: {file_path}")
        try:
            # Intentar leer con pandas primero
            print("  Probando con pandas...")
            df = pd.read_excel(file_path, engine='openpyxl')
            print("  ✔ Lectura exitosa con pandas")
            return df
        except Exception as ex:
            print(f"  ✖ Error con pandas: {str(ex)}")
            logger.error(f"Error leyendo Excel con pandas: {ex}")
            try:
                # Fallback: usar openpyxl directamente
                print("  Probando con openpyxl...")
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active
                
                # Convertir a DataFrame manualmente
                data = []
                headers = None
                
                for row in worksheet.iter_rows(values_only=True):
                    if headers is None:
                        headers = [str(cell).strip() if cell else f'Col_{i}' for i, cell in enumerate(row)]
                    else:
                        data.append(row)
                
                if data:
                    df = pd.DataFrame(data, columns=headers)
                    print("  ✔ Lectura exitosa con openpyxl")
                    return df
                else:
                    print("  ✖ No se encontraron datos en el archivo")
                    return None
                    
            except Exception as e2:
                print(f"  ✖ Error con openpyxl: {str(e2)}")
                logger.error(f"Error leyendo Excel con openpyxl: {e2}")
                return None
    def _get_cell_value(self, row, column_position):
        """Obtener valor de celda por posición"""
        try:
            if column_position < len(row):
                value = row.iloc[column_position]
                if pd.isna(value):
                    return None
                return str(value).strip()
            return None
        except:
            return None        
    def _parse_int_safe(self, value_str):
        """Parsear entero de manera segura"""
        if not value_str:
            return None
        
        try:
            clean_value = str(value_str).strip().replace(',', '')
            return int(float(clean_value))
        except (ValueError, TypeError):
            return None
    def _parse_price_safe(self, price_str):
        """Parsear precio de manera segura"""
        if not price_str:
            return None
        
        try:
            # Limpiar el string
            clean_price = str(price_str).strip()
            clean_price = clean_price.replace('S/.', '').replace('S/', '')
            clean_price = clean_price.replace(',', '')
            clean_price = clean_price.strip()
            
            price = float(clean_price)
            return price if price > 0 else None
            
        except (ValueError, TypeError):
            return None
    def _process_products_fixed(self, df, respuesta):
        """Procesar productos usando posiciones fijas de columnas"""
        print("\n=== INICIANDO PROCESAMIENTO DE PRODUCTOS ===")
        from .models import DetalleRespuestaCotizacion
        from apps.ecommerce.products.models import Product
        
        success_count = 0
        
        print(f"Total de filas a procesar: {len(df)}")
        
        for index, row in df.iterrows():
            try:
                print(f"\nProcesando fila {index + 2}...")
                
                # Obtener valores usando posiciones fijas
                codigo = self._get_cell_value(row, self.OFFICIAL_COLUMNS['CÓDIGO'])
                precio_str = self._get_cell_value(row, self.OFFICIAL_COLUMNS['PRECIO U.'])
                
                print(f"  Código: {codigo}")
                print(f"  Precio (str): {precio_str}")
                
                if not codigo or not precio_str:
                    print("  ✖ Fila omitida: código o precio faltante")
                    continue
                    
                precio_unitario = self._parse_price_safe(precio_str)
                print(f"  Precio (parsed): {precio_unitario}")
                
                if precio_unitario is None or precio_unitario <= 0:
                    error_msg = f"Fila {index + 2}: Precio inválido '{precio_str}'"
                    print(f"  ✖ {error_msg}")
                    self.errors.append(error_msg)
                    continue
                
                # Buscar producto
                producto = None
                try:
                    print(f"  Buscando producto con código: {codigo}")
                    producto = Product.objects.get(code__iexact=str(codigo).strip())
                    print(f"  ✔ Producto encontrado: {producto.name}")
                except Product.DoesNotExist:
                    warning_msg = f"Fila {index + 2}: Producto con código '{codigo}' no encontrado"
                    print(f"  ⚠ {warning_msg}")
                    self.warnings.append(warning_msg)
                    # Continuar sin producto vinculado
                
                # Obtener otros valores
                cantidad_solicitada = self._parse_int_safe(
                    self._get_cell_value(row, self.OFFICIAL_COLUMNS['CANT. SOLICITADA'])
                ) or 1
                
                cantidad_disponible = self._parse_int_safe(
                    self._get_cell_value(row, self.OFFICIAL_COLUMNS['CANT. DISPONIBLE'])
                ) or 0
                
                observaciones = self._get_cell_value(row, self.OFFICIAL_COLUMNS['OBSERVACIONES']) or ""
                
                print(f"  Cant. solicitada: {cantidad_solicitada}")
                print(f"  Cant. disponible: {cantidad_disponible}")
                print(f"  Observaciones: {observaciones}")
                
                # Crear detalle
                print("  Creando detalle de respuesta...")
                DetalleRespuestaCotizacion.objects.create(
                    respuesta=respuesta,
                    producto_code=str(codigo).strip(),
                    producto=producto,
                    precio_unitario=precio_unitario,
                    cantidad_cotizada=cantidad_solicitada,
                    cantidad_disponible=cantidad_disponible,
                    observaciones=observaciones,
                    nombre_producto_proveedor=self._get_cell_value(row, self.OFFICIAL_COLUMNS['PRODUCTO']) or ""
                )
                
                success_count += 1
                print(f"  ✔ Producto procesado exitosamente (Total: {success_count})")
                
            except Exception as e:
                error_msg = f"Fila {index + 2}: Error - {str(e)}"
                print(f"  ✖ {error_msg}")
                self.errors.append(error_msg)
                continue
        
        print(f"\n=== PROCESAMIENTO COMPLETADO ===")
        print(f"Productos procesados exitosamente: {success_count}/{len(df)}")
        print(f"Errores encontrados: {len(self.errors)}")
        print(f"Advertencias: {len(self.warnings)}")
        
        return success_count

    def process_excel_file_with_logging(self):
        """Procesar archivo Excel con logging detallado"""
        from datetime import datetime
        
        print("\n" + "="*50)
        print("INICIANDO PROCESAMIENTO DE ARCHIVO EXCEL")
        print("="*50)
        
        start_time = datetime.now()
        
        try:
            if not self.envio.archivo_respuesta_cliente:
                error_msg = 'No hay archivo para procesar'
                print(f"✖ {error_msg}")
                return {'success': False, 'error': error_msg}
            
            # Leer archivo Excel
            file_path = self.envio.archivo_respuesta_cliente.path
            print(f"\nArchivo a procesar: {file_path}")
            
            validation_result = self._validate_excel_file(file_path)
            
            if not validation_result['valid']:
                print("\n✖ Validación fallida:")
                for error in validation_result['errors']:
                    print(f"  - {error}")
                return {
                    'success': False, 
                    'error': 'Archivo no válido',
                    'details': validation_result['errors'],
                    'validation_errors': validation_result['errors']
                }
            
            # Si la validación es exitosa, procesar
            df = validation_result['dataframe']
            
            # Crear respuesta de cotización
            print("\nCreando respuesta de cotización...")
            respuesta = self._create_cotizacion_response()
            print(f"✔ Respuesta creada con ID: {respuesta.id}")
            
            # Procesar productos
            print("\nProcesando productos...")
            success_count = self._process_products_fixed(df, respuesta)
            
            end_time = datetime.now()
            processing_time = end_time - start_time
            
            print(f"\nTiempo total de procesamiento: {processing_time}")
            
            if success_count > 0:
                print("\n✔ Procesamiento exitoso")
                return {
                    'success': True,
                    'respuesta_id': respuesta.id,
                    'processed_items': success_count,
                    'total_items': len(df),
                    'errors': self.errors,
                    'warnings': self.warnings,
                    'processing_time': str(processing_time)
                }
            else:
                # Si no se procesó ningún producto, eliminar la respuesta creada
                print("\n✖ No se procesaron productos")
                respuesta.delete()
                return {
                    'success': False,
                    'error': 'No se pudo procesar ningún producto',
                    'details': self.errors
                }
                
        except Exception as ex:
            end_time = datetime.now()
            print(f"\n!!! ERROR INESPERADO: {str(ex)}")
            logger.error(f"Error procesando archivo de cotización: {ex}")
            return {
                'success': False,
                'error': f'Error inesperado: {str(ex)}'
            }

    
class EmailCotizacionService:
    """Servicio para envío de emails de cotización"""
    
    @staticmethod
    def send_cotizacion_email(envio, pdf_buffer, excel_buffer):
        """Enviar email de cotización con archivos adjuntos"""
        try:
            from django.conf import settings
            
            supplier = envio.proveedor
            requirement = envio.requerimiento
            
            # Preparar contenido del email
            subject = f"Solicitud de Cotización - {requirement.numero_requerimiento}"
            
            context = {
                'supplier': supplier,
                'requirement': requirement,
                'envio': envio,
                'company_name': getattr(settings, 'COMPANY_NAME', 'Tu Empresa S.A.C.'),
                'company_department': getattr(settings, 'COMPANY_DEPARTMENT', 'Departamento de Compras'),
                'contact_email': getattr(settings, 'CONTACT_EMAIL', 'compras@empresa.com'),
                'contact_phone': getattr(settings, 'CONTACT_PHONE', '+51 999 999 999'),
                'company_address': getattr(settings, 'COMPANY_ADDRESS', 'Av. Principal 123, Lima, Perú'),
                'company_ruc': getattr(settings, 'COMPANY_RUC', '12345678901'),
            }
            
            # Renderizar template de email
            html_content = render_to_string('emails/cotizacion_request.html', context)
            
            # Contenido de texto plano como fallback
            text_content = f"""
Estimado/a {supplier.contact_person},

Saludos cordiales desde el {context['company_department']} de {context['company_name']}.

Nos dirigimos a usted para solicitar una cotización para los productos detallados en los archivos adjuntos.

INFORMACIÓN DEL REQUERIMIENTO:
• Número: {requirement.numero_requerimiento}
• Fecha límite: {requirement.fecha_requerimiento.strftime('%d/%m/%Y')}
• Prioridad: {requirement.prioridad_display}
• Total productos: {requirement.total_productos}

ARCHIVOS ADJUNTOS:
1. PDF con información detallada del requerimiento
2. Excel para completar la cotización (IMPORTANTE: Complete los precios)

INSTRUCCIONES PARA EL PROVEEDOR:
1. Complete los precios unitarios en la columna 'PRECIO U.'
2. Especifique el tiempo de entrega para cada producto
3. Agregue observaciones específicas si es necesario
4. Incluya sus términos de pago y condiciones comerciales
5. Envíe el archivo completado al email: {context['contact_email']}

INFORMACIÓN ADICIONAL A INCLUIR:
• Términos de pago:
• Tiempo de entrega general:
• Validez de la cotización:
• Incluye IGV: SÍ / NO
• Observaciones generales:

Fecha límite para respuesta: {envio.fecha_respuesta_esperada.strftime('%d/%m/%Y')}

Para consultas, contáctenos al teléfono {context['contact_phone']} o responda este email.

Atentamente,
{context['company_department']}
{context['company_name']}
Email: {context['contact_email']}
Teléfono: {context['contact_phone']}

Gracias por su cotización.
            """
            
            # Crear email
            email = EmailMessage(
                subject=subject,
                body=text_content,
                from_email=settings.EMAIL_HOST_USER,
                to=[supplier.email]
            )
            
            # Configurar email HTML
            email.content_subtype = "html"
            email.body = html_content
            
            # Adjuntar PDF
            if pdf_buffer:
                pdf_buffer.seek(0)
                email.attach(
                    f"Requerimiento_{requirement.numero_requerimiento}.pdf",
                    pdf_buffer.getvalue(),
                    'application/pdf'
                )
            
            # Adjuntar Excel (solo uno como solicitaste)
            if excel_buffer:
                excel_buffer.seek(0)
                email.attach(
                    f"Cotizacion_{requirement.numero_requerimiento}.xlsx",
                    excel_buffer.getvalue(),
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            
            # Enviar email
            email.send()
            return True
            
        except Exception as e:
            logger.error(f"Error enviando email de cotización: {e}")
            return False

class CotizacionManager:
    """Manejador principal para operaciones de cotización"""
    
    @staticmethod
    def create_envios_masivos(requirement, supplier_ids, user, fecha_respuesta_esperada=None, notas=None):
        """Crear envíos masivos de cotización"""
        from .models import EnvioCotizacion
        from apps.ecommerce.suppliers.models import Supplier
        from apps.ecommerce.requirements.services import RequirementPDFGenerator
        
        if not fecha_respuesta_esperada:
            fecha_respuesta_esperada = date.today() + timedelta(days=7)
        
        enviados = []
        errores = []
        
        for supplier_id in supplier_ids:
            try:
                supplier = Supplier.objects.get(id=supplier_id, is_active=True)
                
                # Verificar si ya existe envío
                envio_existente = EnvioCotizacion.objects.filter(
                    requerimiento=requirement,
                    proveedor=supplier
                ).first()
                
                if envio_existente:
                    errores.append({
                        'supplier': supplier.company_name,
                        'error': 'Ya existe un envío para este proveedor'
                    })
                    continue
                
                # Determinar método de envío
                metodo_envio = 'email' if supplier.email else 'telefono'
                
                # Crear envío
                envio = EnvioCotizacion.objects.create(
                    requerimiento=requirement,
                    proveedor=supplier,
                    usuario_creacion=user,
                    metodo_envio=metodo_envio,
                    fecha_respuesta_esperada=fecha_respuesta_esperada,
                    notas_envio=notas or '',
                    estado='pendiente'
                )
                
                # Intentar envío automático por email
                if supplier.email:
                    success = CotizacionManager.send_email_cotizacion(envio)
                    if success:
                        envio.estado = 'enviado'
                        envio.fecha_envio = timezone.now()
                        envio.enviado_por = user
                        envio.email_enviado = True
                        envio.fecha_email_enviado = timezone.now()
                        envio.save()
                        
                        enviados.append({
                            'supplier': supplier.company_name,
                            'method': 'email',
                            'contact': supplier.email,
                            'status': 'sent'
                        })
                    else:
                        envio.notas_envio = "Error al enviar email automáticamente"
                        envio.save()
                        errores.append({
                            'supplier': supplier.company_name,
                            'error': 'Error al enviar email'
                        })
                else:
                    # Marcar para envío manual
                    envio.notas_envio = f"Envío manual requerido - Tel: {supplier.phone_primary}"
                    envio.save()
                    
                    enviados.append({
                        'supplier': supplier.company_name,
                        'method': 'manual',
                        'contact': supplier.phone_primary,
                        'status': 'pending_manual'
                    })
                    
            except Supplier.DoesNotExist:
                errores.append({
                    'supplier_id': supplier_id,
                    'error': 'Proveedor no encontrado o inactivo'
                })
            except Exception as e:
                errores.append({
                    'supplier_id': supplier_id,
                    'error': str(e)
                })
        
        return {
            'enviados': enviados,
            'errores': errores,
            'total_enviados': len(enviados),
            'total_errores': len(errores)
        }
    
    @staticmethod
    def send_email_cotizacion(envio):
        """Enviar cotización por email"""
        try:
            from apps.ecommerce.requirements.services import RequirementPDFGenerator
            
            # Generar PDF
            pdf_generator = RequirementPDFGenerator(envio.requerimiento)
            pdf_buffer = pdf_generator.generate_pdf()
            
            # Generar Excel
            excel_generator = RequirementExcelGenerator(envio.requerimiento)
            excel_buffer = excel_generator.create_excel()
            
            # Enviar email
            return EmailCotizacionService.send_cotizacion_email(
                envio, pdf_buffer, excel_buffer
            )
            
        except Exception as e:
            logger.error(f"Error en send_email_cotizacion: {e}")
            return False
    
    @staticmethod
    def confirmar_envio_manual(envio_id, metodo_envio, user, notas=None):
        """Confirmar que se envió manualmente"""
        from .models import EnvioCotizacion
        
        try:
            envio = EnvioCotizacion.objects.get(id=envio_id)
            
            if envio.estado != 'pendiente':
                return False, "Solo se pueden confirmar envíos pendientes"
            
            envio.estado = 'enviado'
            envio.metodo_envio = metodo_envio
            envio.fecha_envio = timezone.now()
            envio.enviado_por = user
            envio.enviado_manualmente = True
            envio.fecha_envio_manual = timezone.now()
            
            if notas:
                envio.notas_envio = f"{envio.notas_envio}\n\nConfirmación manual: {notas}" if envio.notas_envio else f"Confirmación manual: {notas}"
            
            envio.save()
            return True, "Envío confirmado exitosamente"
            
        except EnvioCotizacion.DoesNotExist:
            return False, "Envío no encontrado"
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    def get_cotizacion_analytics(requirement_id=None):
        """Obtener analíticas de cotizaciones"""
        from .models import EnvioCotizacion
        from django.db.models import Count, Q
        
        queryset = EnvioCotizacion.objects.all()
        
        if requirement_id:
            queryset = queryset.filter(requerimiento_id=requirement_id)
        
        # Estadísticas por estado
        stats_by_status = queryset.values('estado').annotate(
            count=Count('id')
        ).order_by('estado')
        
        # Estadísticas por método de envío
        stats_by_method = queryset.values('metodo_envio').annotate(
            count=Count('id')
        ).order_by('metodo_envio')
        
        # Respuestas recibidas
        respuestas_recibidas = queryset.filter(estado='respondido').count()
        
        # Vencidos
        vencidos = queryset.filter(
            Q(estado__in=['pendiente', 'enviado']) &
            Q(fecha_respuesta_esperada__lt=date.today())
        ).count()
        
        return {
            'total_envios': queryset.count(),
            'respuestas_recibidas': respuestas_recibidas,
            'pendientes': queryset.filter(estado='pendiente').count(),
            'enviados': queryset.filter(estado='enviado').count(),
            'vencidos': vencidos,
            'stats_by_status': list(stats_by_status),
            'stats_by_method': list(stats_by_method),
            'tasa_respuesta': (respuestas_recibidas / queryset.count() * 100) if queryset.count() > 0 else 0
        }
        
