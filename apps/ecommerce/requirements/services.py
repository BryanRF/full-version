from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.http import HttpResponse, FileResponse
from django.template.loader import get_template
from rest_framework.decorators import action
from io import BytesIO
import os
import json
from datetime import date
from django.db import transaction
from django.core.exceptions import ValidationError
from .models import Requirement, RequirementDetail
from apps.ecommerce.products.models import Product
class RequirementPDFGenerator:
    """Generador de PDF para requerimientos"""
    
    def __init__(self, requirement):
        self.requirement = requirement
        self.buffer = BytesIO()
        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        self.styles = getSampleStyleSheet()
        self.story = []
        
    def create_custom_styles(self):
        """Crear estilos personalizados"""
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        self.header_style = ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )
        
    def add_header(self):
        """Agregar encabezado del documento"""
        # Título principal
        title = Paragraph("SOLICITUD DE COTIZACIÓN", self.title_style)
        self.story.append(title)
        self.story.append(Spacer(1, 20))
        
        # Información de la empresa (personalizar según tu empresa)
        company_info = f"""
        <b>EMPRESA:</b> Tu Empresa S.A.C.<br/>
        <b>CONTACTO:</b> Departamento de Compras<br/>
        <b>TELÉFONO:</b> +51 999 999 999<br/>
        <b>EMAIL:</b> compras@tuempresa.com<br/>
        <b>FECHA:</b> {date.today().strftime('%d/%m/%Y')}
        """
        
        company_para = Paragraph(company_info, self.normal_style)
        self.story.append(company_para)
        self.story.append(Spacer(1, 20))
        
    def add_requirement_info(self):
        """Agregar información del requerimiento"""
        req_info = f"""
        <b>NÚMERO DE REQUERIMIENTO:</b> {self.requirement.numero_requerimiento}<br/>
        <b>FECHA LÍMITE DE ENTREGA:</b> {self.requirement.fecha_requerimiento.strftime('%d/%m/%Y')}<br/>
        <b>SOLICITANTE:</b> {self.requirement.usuario_solicitante.get_full_name() or self.requirement.usuario_solicitante.username}<br/>
        <b>PRIORIDAD:</b> {self.requirement.prioridad_display}
        """
        
        if self.requirement.notas:
            req_info += f"<br/><b>OBSERVACIONES GENERALES:</b> {self.requirement.notas}"
        
        req_para = Paragraph(req_info, self.normal_style)
        self.story.append(req_para)
        self.story.append(Spacer(1, 20))
        
    def add_products_table(self):
        """Agregar tabla de productos"""
        # Título de la sección
        products_title = Paragraph("PRODUCTOS SOLICITADOS", self.header_style)
        self.story.append(products_title)
        
        # Datos para la tabla
        data = [['ITEM', 'PRODUCTO', 'CANTIDAD', 'UNIDAD', 'OBSERVACIONES']]
        
        for idx, detalle in enumerate(self.requirement.detalles.all(), 1):
            data.append([
                str(idx),
                detalle.producto.name,
                str(detalle.cantidad_solicitada),
                detalle.unidad_medida,
                detalle.observaciones or '-'
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[0.8*inch, 3*inch, 1*inch, 1*inch, 2*inch])
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Contenido
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Producto alineado a la izquierda
            ('ALIGN', (4, 1), (4, -1), 'LEFT'),  # Observaciones alineadas a la izquierda
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        self.story.append(table)
        self.story.append(Spacer(1, 30))
        
    def add_footer_instructions(self):
        """Agregar instrucciones para el proveedor"""
        instructions = """
        <b>INSTRUCCIONES PARA EL PROVEEDOR:</b><br/>
        <br/>
        1. Por favor, proporcionar cotización detallada incluyendo precios unitarios y totales.<br/>
        2. Especificar tiempo de entrega para cada producto.<br/>
        3. Incluir condiciones de pago y garantías.<br/>
        4. Enviar cotización en formato PDF a: compras@tuempresa.com<br/>
        5. Referenciar el número de requerimiento en toda comunicación.<br/>
        <br/>
        <b>FECHA LÍMITE PARA COTIZACIÓN:</b> 3 días hábiles a partir de la fecha de este documento.<br/>
        <br/>
        Gracias por su colaboración.
        """
        
        instructions_para = Paragraph(instructions, self.normal_style)
        self.story.append(instructions_para)
        
    def generate_pdf(self):
        """Generar el PDF completo"""
        self.create_custom_styles()
        self.add_header()
        self.add_requirement_info()
        self.add_products_table()
        self.add_footer_instructions()
        
        # Construir el documento
        self.doc.build(self.story)
        
        # Retornar el buffer
        self.buffer.seek(0)
        return self.buffer



import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import date, timedelta

class RequirementExcelGenerator:
    """Generador de Excel para requerimientos"""
    
    def __init__(self, requirement):
        self.requirement = requirement
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Requerimiento"
        
    def create_excel(self):
        """Crear archivo Excel para requerimiento"""
        self._setup_styles()
        self._add_products_table()
        self._adjust_columns()
        
        # Guardar en buffer
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _setup_styles(self):
        """Configurar estilos del Excel"""
        # Estilos para encabezados
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilos para información
        self.info_font = Font(bold=True, size=11)
        self.normal_font = Font(size=10)
        
        # Bordes
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    

    def _add_products_table(self):
        """Agregar tabla de productos"""
        start_row = 1
        
        # Encabezados de la tabla simplificados
        headers =  [             'CÓDIGO', 'PRODUCTO', 'CATEGORÍA', 'CANT. SOLICITADA', 
                    'UNIDAD', 'CANT. DISPONIBLE', 'PRECIO U.', 'PRECIO TOTAL', 'OBSERVACIONES'         ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Escribir datos de productos
        current_row = start_row + 1
        for detalle in self.requirement.detalles.all():
            row_data = [
                detalle.producto.code,
                detalle.producto.name,
                detalle.producto.category.name if detalle.producto.category else "Sin categoría",
                detalle.cantidad_solicitada,
                detalle.unidad_medida,
                "",
                "",  # Precio unitario - campo vacío para completar
                "",  # Precio total - se calculará automáticamente
                detalle.observaciones or ""
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=current_row, column=col)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                
                # Alineación especial para algunas columnas
                if col in [4, 6, 7, 8]: 
                    cell.alignment = Alignment(horizontal="center")
                elif col == 1:  # Código
                    cell.alignment = Alignment(horizontal="center")
                
                # Color de fondo para campos editables
                if col in [7,6,9]:  # Precio unitario
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
         
            current_row += 1
        
        # Agregar fórmulas para precio total
        last_row = current_row - 1
        for row in range(start_row + 1, current_row):
            formula_cell = self.worksheet.cell(row=row, column=8)  # Columna H (Precio Total)
            formula_cell.value = f"=F{row}*G{row}"
            formula_cell.number_format = '#,##0.00'
            formula_cell.border = self.border
            formula_cell.alignment = Alignment(horizontal="center")
        return current_row + 1
    def _adjust_columns(self):
        """Ajustar ancho de columnas automáticamente basado en el contenido"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Ajustar el ancho basado en el contenido más largo de cada columna
            for cell in column:
                try:
                    # Calcular longitud del contenido
                    if cell.value:
                        cell_length = len(str(cell.value))
                        # Considerar saltos de línea en el cálculo
                        if '\n' in str(cell.value):
                            cell_length = max(len(line) for line in str(cell.value).split('\n'))
                        
                        # Ajustar para encabezados con formato especial
                        if cell.font.bold:
                            cell_length *= 1.2
                            
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Establecer un ancho mínimo y máximo razonable
            min_width = 8  # Ancho mínimo para que se vea bien
            max_width = 50  # Ancho máximo para evitar columnas demasiado anchas
            
            # Ajustes específicos para ciertas columnas
            if column_letter == 'A':  # Código
                adjusted_width = max(min_width, min(max_width, max_length + 2))
            elif column_letter == 'B':  # Producto
                adjusted_width = max(min_width, min(40, max_length + 2))  # Máximo 40 para nombres largos
            elif column_letter == 'I':  # Observaciones
                adjusted_width = max(min_width, min(30, max_length + 2))  # Máximo 30 para observaciones
            else:
                adjusted_width = max(min_width, min(max_width, max_length + 2))
            
            self.worksheet.column_dimensions[column_letter].width = adjusted_width


# También agregar al final del archivo existing RequirementPDFGenerator la función para generar Excel de cotización:


class ComparisonExcelGenerator:
    """Generador de Excel para comparación de cotizaciones"""
    
    def __init__(self, requirement):
        self.requirement = requirement
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Comparación Cotizaciones"
        
    def create_comparison_excel(self):
        """Crear archivo Excel de comparación"""
        self._setup_styles()
        self._add_header_info()
        self._add_comparison_table()
        self._add_detailed_analysis()
        self._adjust_columns()
        
        # Guardar en buffer
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _setup_styles(self):
        """Configurar estilos del Excel"""
        # Estilos para encabezados
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilos para información
        self.info_font = Font(bold=True, size=11)
        self.normal_font = Font(size=10)
        self.title_font = Font(bold=True, size=14, color="366092")
        
        # Bordes
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def _add_header_info(self):
        """Agregar información del requerimiento"""
        from django.conf import settings
        
        # Título principal
        self.worksheet.merge_cells('A1:H1')
        title_cell = self.worksheet['A1']
        title_cell.value = "COMPARACIÓN DE COTIZACIONES"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment
        
        # Información de la empresa
        row = 3
        company_info = [
            ("Empresa:", getattr(settings, 'COMPANY_NAME', 'Tu Empresa S.A.C.')),
            ("Departamento:", getattr(settings, 'COMPANY_DEPARTMENT', 'Departamento de Compras')),
            ("Requerimiento:", self.requirement.numero_requerimiento),
            ("Fecha Análisis:", date.today().strftime('%d/%m/%Y')),
        ]
        
        for label, value in company_info:
            self.worksheet[f'A{row}'] = label
            self.worksheet[f'A{row}'].font = self.info_font
            self.worksheet[f'B{row}'] = value
            self.worksheet[f'B{row}'].font = self.normal_font
            row += 1
        
        return row + 2
    
    def _add_comparison_table(self):
        """Agregar tabla de comparación"""
        from apps.ecommerce.cotizacion.models import RespuestaCotizacion
        
        start_row = 8
        
        # Obtener respuestas de cotización
        respuestas = RespuestaCotizacion.objects.filter(
            envio__requerimiento=self.requirement
        ).prefetch_related('detalles', 'envio__proveedor')
        
        # Título de la sección
        self.worksheet.merge_cells(f'A{start_row}:H{start_row}')
        section_title = self.worksheet[f'A{start_row}']
        section_title.value = "RESUMEN COMPARATIVO"
        section_title.font = self.info_font
        section_title.alignment = self.center_alignment
        
        # Encabezados de la tabla
        headers = [
            'PROVEEDOR', 'TOTAL COTIZADO', 'PRODUCTOS', 'COBERTURA %', 
            'TÉRMINOS PAGO', 'TIEMPO ENTREGA', 'INCLUYE IGV', 'CALIFICACIÓN'
        ]
        
        header_row = start_row + 2
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=header_row, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Datos de comparación
        current_row = header_row + 1
        total_productos = self.requirement.total_productos
        
        for respuesta in respuestas:
            productos_cotizados = respuesta.detalles.count()
            cobertura = round((productos_cotizados / total_productos) * 100) if total_productos > 0 else 0
            
            row_data = [
                respuesta.envio.proveedor.company_name,
                float(respuesta.total_cotizado),
                productos_cotizados,
                cobertura,
                respuesta.terminos_pago or 'No especificado',
                respuesta.tiempo_entrega or 'No especificado',
                'SÍ' if respuesta.incluye_igv else 'NO',
                respuesta.envio.proveedor.rating
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=current_row, column=col)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                
                # Formato especial para algunas columnas
                if col == 2:  # Total cotizado
                    cell.number_format = '"S/." #,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col in [3, 4, 8]:  # Productos, cobertura, calificación
                    cell.alignment = self.center_alignment
                elif col == 4:  # Cobertura %
                    cell.number_format = '0"%"'
            
            current_row += 1
        
        return current_row + 2
    
    def _add_detailed_analysis(self):
        """Agregar análisis detallado por producto"""
        from apps.ecommerce.cotizacion.models import RespuestaCotizacion, DetalleRespuestaCotizacion
        
        start_row = self.worksheet.max_row + 2
        
        # Título de la sección
        self.worksheet.merge_cells(f'A{start_row}:I{start_row}')
        section_title = self.worksheet[f'A{start_row}']
        section_title.value = "ANÁLISIS DETALLADO POR PRODUCTO"
        section_title.font = self.info_font
        section_title.alignment = self.center_alignment
        
        # Obtener todos los productos del requerimiento
        productos_requerimiento = {}
        for detalle in self.requirement.detalles.all():
            productos_requerimiento[detalle.producto.code] = {
                'codigo': detalle.producto.code,
                'nombre': detalle.producto.name,
                'cantidad_solicitada': detalle.cantidad_solicitada,
                'cotizaciones': []
            }
        
        # Obtener cotizaciones para cada producto
        detalles_respuesta = DetalleRespuestaCotizacion.objects.filter(
            respuesta__envio__requerimiento=self.requirement
        ).select_related('respuesta__envio__proveedor', 'producto')
        
        for detalle in detalles_respuesta:
            if detalle.producto_code in productos_requerimiento:
                productos_requerimiento[detalle.producto_code]['cotizaciones'].append({
                    'proveedor': detalle.respuesta.envio.proveedor.company_name,
                    'precio_unitario': float(detalle.precio_unitario),
                    'cantidad_disponible': detalle.cantidad_disponible,
                    'tiempo_entrega': detalle.tiempo_entrega_especifico
                })
        
        # Renderizar análisis por producto
        current_row = start_row + 2
        
        for codigo, producto_data in productos_requerimiento.items():
            # Encabezado del producto
            self.worksheet.merge_cells(f'A{current_row}:I{current_row}')
            product_header = self.worksheet[f'A{current_row}']
            product_header.value = f"{codigo} - {producto_data['nombre']}"
            product_header.font = Font(bold=True, size=11, color="000080")
            product_header.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            product_header.border = self.border
            
            current_row += 1
            
            if producto_data['cotizaciones']:
                # Encabezados de la sub-tabla
                sub_headers = [
                    'PROVEEDOR', 'PRECIO UNIT.', 'CANTIDAD DISP.', 'TIEMPO ENTREGA', 
                    'DIFERENCIA %', 'RANKING', '', '', ''
                ]
                
                for col, header in enumerate(sub_headers, 1):
                    if col <= 6:  # Solo las primeras 6 columnas
                        cell = self.worksheet.cell(row=current_row, column=col)
                        cell.value = header
                        cell.font = Font(bold=True, size=9)
                        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                        cell.border = self.border
                        cell.alignment = self.center_alignment
                
                current_row += 1
                
                # Ordenar cotizaciones por precio
                cotizaciones_ordenadas = sorted(
                    producto_data['cotizaciones'], 
                    key=lambda x: x['precio_unitario']
                )
                
                precio_menor = cotizaciones_ordenadas[0]['precio_unitario'] if cotizaciones_ordenadas else 0
                
                for rank, cot in enumerate(cotizaciones_ordenadas, 1):
                    diferencia = ((cot['precio_unitario'] - precio_menor) / precio_menor * 100) if precio_menor > 0 else 0
                    
                    sub_row_data = [
                        cot['proveedor'],
                        cot['precio_unitario'],
                        cot['cantidad_disponible'],
                        cot['tiempo_entrega'] or 'No especificado',
                        diferencia,
                        rank
                    ]
                    
                    for col, value in enumerate(sub_row_data, 1):
                        cell = self.worksheet.cell(row=current_row, column=col)
                        cell.value = value
                        cell.font = Font(size=9)
                        cell.border = self.border
                        
                        # Formato especial
                        if col == 2:  # Precio unitario
                            cell.number_format = '"S/." #,##0.00'
                            cell.alignment = Alignment(horizontal="right")
                        elif col in [3, 6]:  # Cantidad y ranking
                            cell.alignment = self.center_alignment
                        elif col == 5:  # Diferencia %
                            cell.number_format = '0.0"%"'
                            cell.alignment = Alignment(horizontal="right")
                            if rank == 1:
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    
                    current_row += 1
            else:
                # Sin cotizaciones
                self.worksheet.merge_cells(f'A{current_row}:F{current_row}')
                no_quotes = self.worksheet[f'A{current_row}']
                no_quotes.value = "No se recibieron cotizaciones para este producto"
                no_quotes.font = Font(italic=True, color="FF0000")
                no_quotes.alignment = self.center_alignment
                current_row += 1
            
            current_row += 1  # Espacio entre productos
    
    def _adjust_columns(self):
        """Ajustar ancho de columnas"""
        column_widths = {
            'A': 25,  # Proveedor/Código
            'B': 15,  # Total/Precio
            'C': 12,  # Productos/Cantidad
            'D': 12,  # Cobertura/Tiempo
            'E': 20,  # Términos/Diferencia
            'F': 15,  # Tiempo/Ranking
            'G': 12,  # IGV
            'H': 12,  # Calificación
            'I': 15   # Extra
        }
        
        for column, width in column_widths.items():
            self.worksheet.column_dimensions[column].width = width


class ComparisonPDFGenerator:
    """Generador de PDF para comparación de cotizaciones"""
    
    def __init__(self, requirement):
        self.requirement = requirement
        self.buffer = BytesIO()
        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=A4,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        self.styles = getSampleStyleSheet()
        self.story = []
        
    def create_comparison_pdf(self):
        """Crear PDF de comparación"""
        self._create_custom_styles()
        self._add_header()
        self._add_summary_table()
        self._add_analysis()
        self._add_recommendations()
        
        # Construir el documento
        self.doc.build(self.story)
        self.buffer.seek(0)
        return self.buffer
    
    def _create_custom_styles(self):
        """Crear estilos personalizados"""
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        self.header_style = ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        self.normal_style = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            spaceAfter=6
        )
    
    def _add_header(self):
        """Agregar encabezado del documento"""
        from django.conf import settings
        
        # Título principal
        title = Paragraph("COMPARACIÓN DE COTIZACIONES", self.title_style)
        self.story.append(title)
        self.story.append(Spacer(1, 20))
        
        # Información de la empresa y requerimiento
        company_info = f"""
        <b>EMPRESA:</b> {getattr(settings, 'COMPANY_NAME', 'Tu Empresa S.A.C.')}<br/>
        <b>DEPARTAMENTO:</b> {getattr(settings, 'COMPANY_DEPARTMENT', 'Departamento de Compras')}<br/>
        <b>REQUERIMIENTO:</b> {self.requirement.numero_requerimiento}<br/>
        <b>FECHA ANÁLISIS:</b> {date.today().strftime('%d/%m/%Y')}<br/>
        <b>PRIORIDAD:</b> {self.requirement.prioridad_display}
        """
        
        info_para = Paragraph(company_info, self.normal_style)
        self.story.append(info_para)
        self.story.append(Spacer(1, 20))
    
    def _add_summary_table(self):
        """Agregar tabla resumen de comparación"""
        from apps.ecommerce.cotizacion.models import RespuestaCotizacion
        
        # Título de la sección
        summary_title = Paragraph("RESUMEN COMPARATIVO", self.header_style)
        self.story.append(summary_title)
        
        # Obtener respuestas
        respuestas = RespuestaCotizacion.objects.filter(
            envio__requerimiento=self.requirement
        ).prefetch_related('detalles', 'envio__proveedor')
        
        if not respuestas.exists():
            no_data = Paragraph("No se encontraron cotizaciones para comparar.", self.normal_style)
            self.story.append(no_data)
            return
        
        # Datos para la tabla
        data = [['Proveedor', 'Total', 'Productos', 'Términos Pago', 'Tiempo Entrega']]
        
        for respuesta in respuestas:
            data.append([
                respuesta.envio.proveedor.company_name,
                f"S/ {respuesta.total_cotizado:,.2f}",
                str(respuesta.detalles.count()),
                respuesta.terminos_pago or 'No especificado',
                respuesta.tiempo_entrega or 'No especificado'
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[2*inch, 1.2*inch, 0.8*inch, 1.5*inch, 1.5*inch])
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Contenido
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),  # Total alineado a la derecha
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        self.story.append(table)
        self.story.append(Spacer(1, 30))
    
    def _add_analysis(self):
        """Agregar análisis de datos"""
        from apps.ecommerce.cotizacion.models import RespuestaCotizacion
        
        # Título de la sección
        analysis_title = Paragraph("ANÁLISIS DE COTIZACIONES", self.header_style)
        self.story.append(analysis_title)
        
        respuestas = RespuestaCotizacion.objects.filter(
            envio__requerimiento=self.requirement
        )
        
        if respuestas.exists():
            precios = [float(r.total_cotizado) for r in respuestas]
            
            analysis_text = f"""
            <b>ESTADÍSTICAS:</b><br/>
            • Total de cotizaciones recibidas: {len(precios)}<br/>
            • Precio menor: S/ {min(precios):,.2f}<br/>
            • Precio mayor: S/ {max(precios):,.2f}<br/>
            • Precio promedio: S/ {sum(precios)/len(precios):,.2f}<br/>
            • Ahorro potencial: S/ {max(precios) - min(precios):,.2f}<br/>
            • Diferencia porcentual: {((max(precios) - min(precios)) / max(precios) * 100):.1f}%
            """
            
            analysis_para = Paragraph(analysis_text, self.normal_style)
            self.story.append(analysis_para)
            self.story.append(Spacer(1, 20))
    
    def _add_recommendations(self):
        """Agregar recomendaciones"""
        # Título de la sección
        rec_title = Paragraph("RECOMENDACIONES", self.header_style)
        self.story.append(rec_title)
        
        recommendations = """
        <b>CRITERIOS DE EVALUACIÓN SUGERIDOS:</b><br/>
        <br/>
        1. <b>Precio:</b> Evaluar la oferta económica más competitiva.<br/>
        2. <b>Calidad:</b> Considerar la reputación y calificación del proveedor.<br/>
        3. <b>Tiempo de entrega:</b> Verificar cumplimiento de plazos requeridos.<br/>
        4. <b>Términos de pago:</b> Analizar condiciones financieras favorables.<br/>
        5. <b>Cobertura:</b> Priorizar proveedores que coticen todos los productos.<br/>
        <br/>
        <b>PRÓXIMOS PASOS:</b><br/>
        • Verificar referencias de proveedores seleccionados<br/>
        • Negociar términos finales si es necesario<br/>
        • Formalizar la orden de compra<br/>
        • Establecer cronograma de entrega
        """
        
        rec_para = Paragraph(recommendations, self.normal_style)
        self.story.append(rec_para)       
        
        
        
class RequirementUpdateService:
    """Servicio para manejar la actualización completa de requerimientos"""
    
    def __init__(self, requirement):
        self.requirement = requirement
        
    def can_update(self):
        """Verificar si el requerimiento puede ser actualizado"""
        return self.requirement.can_edit
    
    def validate_update_data(self, data):
        """Validar datos de actualización"""
        errors = []
        
        # Validar fecha requerimiento
        fecha_requerimiento = data.get('fecha_requerimiento')
        if fecha_requerimiento:
            from datetime import date
            if fecha_requerimiento < date.today():
                errors.append("La fecha requerida no puede ser en el pasado")
        
        # Validar detalles si se proporcionan
        detalles = data.get('detalles')
        if detalles is not None:
            if isinstance(detalles, str):
                try:
                    detalles = json.loads(detalles)
                except json.JSONDecodeError:
                    errors.append("Formato JSON inválido en detalles")
                    return errors
            
            if not detalles or len(detalles) == 0:
                errors.append("Debe incluir al menos un producto")
            
            # Validar productos únicos
            productos_ids = []
            for detalle in detalles:
                producto_id = detalle.get('producto_id')
                if not producto_id:
                    errors.append("producto_id es requerido en cada detalle")
                    continue
                    
                if producto_id in productos_ids:
                    errors.append("No se pueden repetir productos en el mismo requerimiento")
                productos_ids.append(producto_id)
                
                # Validar cantidad
                cantidad = detalle.get('cantidad_solicitada', 1)
                if cantidad <= 0:
                    errors.append("La cantidad solicitada debe ser mayor a 0")
        
        return errors
    
    @transaction.atomic
    def update_requirement(self, validated_data):
        """Actualizar requerimiento completo con detalles"""
        
        if not self.can_update():
            raise ValidationError("Este requerimiento no puede ser editado en su estado actual")
        
        # Extraer detalles del validated_data
        detalles_data = validated_data.pop('detalles', None)
        
        # Actualizar campos básicos del requerimiento
        for field, value in validated_data.items():
            setattr(self.requirement, field, value)
        
        self.requirement.save()
        
        # Si se proporcionaron detalles, actualizar completamente
        if detalles_data is not None:
            self._update_requirement_details(detalles_data)
        
        return self.requirement
    
    def _update_requirement_details(self, detalles_data):
        """Actualizar detalles del requerimiento"""
        
        # Parsear detalles si viene como string JSON
        if isinstance(detalles_data, str):
            detalles_data = json.loads(detalles_data)
        
        # Obtener IDs de productos actuales
        current_detail_ids = set(
            self.requirement.detalles.values_list('id', flat=True)
        )
        
        # Obtener IDs de productos en la actualización
        updated_detail_ids = set()
        
        for detalle_data in detalles_data:
            detalle_id = detalle_data.get('id')
            producto_id = detalle_data.get('producto_id')
            
            # Verificar que el producto existe
            try:
                producto = Product.objects.get(id=producto_id)
            except Product.DoesNotExist:
                raise ValidationError(f"Producto con ID {producto_id} no existe")
            
            if detalle_id:
                # Actualizar detalle existente
                try:
                    detalle = RequirementDetail.objects.get(
                        id=detalle_id, 
                        requerimiento=self.requirement
                    )
                    detalle.producto = producto
                    detalle.cantidad_solicitada = detalle_data.get('cantidad_solicitada', 1)
                    detalle.unidad_medida = detalle_data.get('unidad_medida', 'unidad')
                    detalle.observaciones = detalle_data.get('observaciones', '')
                    detalle.save()
                    updated_detail_ids.add(detalle_id)
                except RequirementDetail.DoesNotExist:
                    # Si no existe, crear nuevo
                    detalle = RequirementDetail.objects.create(
                        requerimiento=self.requirement,
                        producto=producto,
                        cantidad_solicitada=detalle_data.get('cantidad_solicitada', 1),
                        unidad_medida=detalle_data.get('unidad_medida', 'unidad'),
                        observaciones=detalle_data.get('observaciones', '')
                    )
                    updated_detail_ids.add(detalle.id)
            else:
                # Crear nuevo detalle
                detalle = RequirementDetail.objects.create(
                    requerimiento=self.requirement,
                    producto=producto,
                    cantidad_solicitada=detalle_data.get('cantidad_solicitada', 1),
                    unidad_medida=detalle_data.get('unidad_medida', 'unidad'),
                    observaciones=detalle_data.get('observaciones', '')
                )
                updated_detail_ids.add(detalle.id)
        
        # Eliminar detalles que ya no están en la actualización
        details_to_delete = current_detail_ids - updated_detail_ids
        if details_to_delete:
            RequirementDetail.objects.filter(
                id__in=details_to_delete,
                requerimiento=self.requirement
            ).delete()
    
    def add_product(self, producto_id, cantidad_solicitada=1, unidad_medida='unidad', observaciones=''):
        """Agregar un producto al requerimiento"""
        
        if not self.can_update():
            raise ValidationError("Este requerimiento no puede ser editado")
        
        try:
            producto = Product.objects.get(id=producto_id)
        except Product.DoesNotExist:
            raise ValidationError(f"Producto con ID {producto_id} no existe")
        
        # Verificar que el producto no esté ya en el requerimiento
        if self.requirement.detalles.filter(producto=producto).exists():
            raise ValidationError("Este producto ya está en el requerimiento")
        
        detalle = RequirementDetail.objects.create(
            requerimiento=self.requirement,
            producto=producto,
            cantidad_solicitada=cantidad_solicitada,
            unidad_medida=unidad_medida,
            observaciones=observaciones
        )
        
        return detalle
    
    def remove_product(self, producto_id):
        """Remover un producto del requerimiento"""
        
        if not self.can_update():
            raise ValidationError("Este requerimiento no puede ser editado")
        
        try:
            detalle = self.requirement.detalles.get(producto_id=producto_id)
            detalle.delete()
            return True
        except RequirementDetail.DoesNotExist:
            raise ValidationError("El producto no está en este requerimiento")
    
    def update_product_quantity(self, producto_id, nueva_cantidad):
        """Actualizar cantidad de un producto específico"""
        
        if not self.can_update():
            raise ValidationError("Este requerimiento no puede ser editado")
        
        if nueva_cantidad <= 0:
            raise ValidationError("La cantidad debe ser mayor a 0")
        
        try:
            detalle = self.requirement.detalles.get(producto_id=producto_id)
            detalle.cantidad_solicitada = nueva_cantidad
            detalle.save()
            return detalle
        except RequirementDetail.DoesNotExist:
            raise ValidationError("El producto no está en este requerimiento")


class RequirementComparisonService:
    """Servicio para comparar versiones de requerimientos"""
    
    @staticmethod
    def compare_requirements(requirement1, requirement2):
        """Comparar dos requerimientos y retornar diferencias"""
        
        differences = {
            'basic_fields': {},
            'products': {
                'added': [],
                'removed': [],
                'modified': []
            }
        }
        
        # Comparar campos básicos
        basic_fields = [
            'fecha_requerimiento', 'prioridad', 'estado', 'notas'
        ]
        
        for field in basic_fields:
            value1 = getattr(requirement1, field, None)
            value2 = getattr(requirement2, field, None)
            
            if value1 != value2:
                differences['basic_fields'][field] = {
                    'from': value1,
                    'to': value2
                }
        
        # Comparar productos
        products1 = {d.producto_id: d for d in requirement1.detalles.all()}
        products2 = {d.producto_id: d for d in requirement2.detalles.all()}
        
        # Productos agregados
        added_product_ids = set(products2.keys()) - set(products1.keys())
        for product_id in added_product_ids:
            detalle = products2[product_id]
            differences['products']['added'].append({
                'producto_id': product_id,
                'producto_name': detalle.producto.name,
                'cantidad_solicitada': detalle.cantidad_solicitada,
                'unidad_medida': detalle.unidad_medida
            })
        
        # Productos removidos
        removed_product_ids = set(products1.keys()) - set(products2.keys())
        for product_id in removed_product_ids:
            detalle = products1[product_id]
            differences['products']['removed'].append({
                'producto_id': product_id,
                'producto_name': detalle.producto.name,
                'cantidad_solicitada': detalle.cantidad_solicitada,
                'unidad_medida': detalle.unidad_medida
            })
        
        # Productos modificados
        common_product_ids = set(products1.keys()) & set(products2.keys())
        for product_id in common_product_ids:
            detalle1 = products1[product_id]
            detalle2 = products2[product_id]
            
            if (detalle1.cantidad_solicitada != detalle2.cantidad_solicitada or
                detalle1.unidad_medida != detalle2.unidad_medida or
                detalle1.observaciones != detalle2.observaciones):
                
                differences['products']['modified'].append({
                    'producto_id': product_id,
                    'producto_name': detalle1.producto.name,
                    'changes': {
                        'cantidad_solicitada': {
                            'from': detalle1.cantidad_solicitada,
                            'to': detalle2.cantidad_solicitada
                        } if detalle1.cantidad_solicitada != detalle2.cantidad_solicitada else None,
                        'unidad_medida': {
                            'from': detalle1.unidad_medida,
                            'to': detalle2.unidad_medida
                        } if detalle1.unidad_medida != detalle2.unidad_medida else None,
                        'observaciones': {
                            'from': detalle1.observaciones,
                            'to': detalle2.observaciones
                        } if detalle1.observaciones != detalle2.observaciones else None
                    }
                })
        
        return differences
    
    @staticmethod
    def has_significant_changes(differences):
        """Determinar si hay cambios significativos"""
        return (
            len(differences['basic_fields']) > 0 or
            len(differences['products']['added']) > 0 or
            len(differences['products']['removed']) > 0 or
            len(differences['products']['modified']) > 0
        )